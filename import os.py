#ссылка на диск с нужными файлами https://disk.yandex.ru/d/Bt_XUn8XoK4VHw
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime, timedelta
import re
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
# Массив для сравнения числовых значениq
d = ['B565ОО799', 'B587OO799', 'TeltonikaGH5200', 'X403BB799(22)', 'Е753НО799', 'Е877УК799', 'Е886НО799', 'К376ХС799', 'К403ХС799', 'К729ЕТ790', 'М564ХМ799', 'О520УЕ799', 'О645УЕ799', 'Р287АН977', 'Р292АН977', 'Р309АН977', 'Р718РК797', 'Т033ВВ977', 'Т046ВВ977', 'Т674НА797', 'Т872ВК799', 'У675КМ797', 'У720КМ797', 'Х464ВВ799', 'Х576ВВ799', '№1 Генератор ДЭУ', '№2 Генератор ДЭУ', '№3 Генератор ДЭУ', '№4 Генератор ДЭУ', '№5 Генератор Хютер', '№7 Генератор ДЭУ', '№8 Генератор ДЭУ', '№9 Генератор ДЭУ', '№11 Генератор ДЭУ']
a = 'ABCDEFGHIGKLMNOPQRSTUVWXYZ'

def clean_time_string(input_string):
    # Паттерн для времени в формате чч:мм:сс
    time_pattern = r'^\d{2}:\d{2}:\d{2}$'

    # Если строка начинается с времени в формате чч:мм:сс, возвращаем ее без изменений
    if re.match(time_pattern, input_string):
        return input_string
    else:
        # Иначе убираем все, кроме времени
        match = re.search(r'(\d{2}:\d{2}:\d{2})', input_string)
        if match:
            return match.group(0)
        else:
            return None 
def compare_times(time1, time2):
    # Преобразование строковых значений времени в объекты datetime.time
    time1_obj = datetime.strptime(time1, '%H:%M:%S').time()
    time2_obj = datetime.strptime(time2, '%H:%M:%S').time()

    # Разница между временами
    time_diff = datetime.combine(datetime.date.today(), time1_obj) - datetime.combine(datetime.date.today(), time2_obj)
    
    # Проверка условия: разница больше часа
    if abs(time_diff) > timedelta(hours=1):
        print("Разница между временами больше часа")
    else:
        print("Разница между временами меньше или равна часу")
def cgc(time_str):
    hours, minutes, seconds = map(int, time_str.split(':'))
    time_delta = timedelta(hours=hours, minutes=minutes, seconds=seconds)
    return time_delta
def convert_to_time(time_str):
    r =  datetime.strptime(time_str, '%H:%M:%S')
    return r.strftime('%H:%M:%S')
def cgc(time_str):

    hours, minutes, seconds = map(int, time_str.split(':'))

    # Создаем объект timedelta
    time_delta = timedelta(hours=hours, minutes=minutes, seconds=seconds)
    return time_delta

def extract_time_from_datetime(datetime_str):
    try:
        datetime_obj = datetime.strptime(datetime_str, '%d.%m.%Y %H:%M:%S')
        return datetime_obj.strftime('23:59:59')
    except ValueError:
        return datetime_str
def check_and_remove_prefix(program):
    if program.startswith('~$'):
        return program[2:]  # Удаляем первые два символа
    else:
        return program
def convert_date(date_str, full_date_str):
    day, month = date_str.split('.')
    _, _, year = full_date_str.split('.')
    return f"{year}-{month}-{day}"
def next_day(input_date):
    date_format = "%Y-%m-%d"
    input_datetime = datetime.strptime(input_date, date_format)
    next_date = input_datetime + timedelta(days=1)
    return next_date.strftime(date_format)
def next_dayy(input_date):
    date_format = "%Y-%m-%d"
    input_datetime = datetime.strptime(input_date, date_format)
    next_date = input_datetime + timedelta(days=1)
    return next_date.day


mgss = "C:/Users/Никита/Downloads/Машины МГСС.xlsx"

desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
wb = Workbook()
for sheet_name in d:
    wb.create_sheet(title=sheet_name)
file_path = os.path.join(desktop_path, 'тест.xlsx')
wb.save(file_path)
print(f'Файл Excel создан и сохранен по пути: {file_path}')

mgss_file = openpyxl.load_workbook(mgss)
mgss_sheetnames = mgss_file.sheetnames
wb_sheetnames = wb.sheetnames

download_folder = os.path.expanduser("~/Downloads")
files = os.listdir(download_folder)
    # Сортируем файлы по времени последнего изменения
files.sort(key=lambda x: os.path.getmtime(os.path.join(download_folder, x)), reverse=True)
        # Находим первый файл (последний скачанный)    
program = files[0]
latest_file = check_and_remove_prefix(program)
# Открыть последний скачанный файл
download_path = os.path.join(os.path.expanduser('~'), 'Downloads')
file = os.path.join(download_path, latest_file)
print(latest_file)

print(file)
downl_file = openpyxl.load_workbook(file)
downl_sheet = downl_file["Поездки"]
column = downl_sheet.max_column
raww = downl_sheet.max_row
# Создать новую книгу и лист для отфильтрованных данных
wb = openpyxl.load_workbook(file_path)
filtered_sheet = wb["B565ОО799"]

# Фильтрация данных и копирование в новый лист
for row in downl_sheet.iter_rows(values_only=True):
    if len(str(row[0])) in [1, 2]:  # Проверяем длину значения в ячейке A
        filtered_sheet.append(row)
        
columns_to_delete = filtered_sheet.iter_cols(min_col=9, max_col=filtered_sheet.max_column)

# Удаление столбцов
for column in reversed(list(columns_to_delete)):
    filtered_sheet.delete_cols(column[0].column, len(column))
# Сохранить
# новую книгу с отфильтрованными данными

wb.save(file_path)

print("Данные успешно скопированы и сохранены в файле", file_path)


for c in range(len(mgss_sheetnames)):
    sheet_name = mgss_sheetnames[c]
    filtered_sheet = mgss_file[sheet_name]
    max_column = filtered_sheet.max_column
    max_raw = filtered_sheet.max_row
    

    # Перебираем ячейки в строке 1
    for column in range(1, max_column + 1):
        cell_value = filtered_sheet.cell(row=1, column=column).value
        if cell_value == "Тип тс":
            da = a[int(column - 1)]
            v = column 
    print(column)
    
    for n in range(1, max_raw + 1):
        clet = da  + str(n)
        cell1 = filtered_sheet[clet].value
        for i in range(len(d)):
                if cell1 == d[i]:
                    v = d[i]
                    sheet2 = wb[v]
                    for row in filtered_sheet.iter_rows(min_row=n, max_row=n, min_col=1, max_col=column):
                        sheet2.append([cell.value for cell in row])
wb.save(file_path)
mgss_file.close()





# Открываем файл Excel

sheet = wb['B565ОО799']

date_input = "13.02"
full_date_input = "01.01.2024"
converted_date = convert_date(date_input, full_date_input)
print(converted_date)


# Перебираем ячейки во втором столбце
for cell in sheet['B']:
    if 'по' in str(cell.value):
        
        date1, date2 = cell.value.split(' по ')
        date1 = date1.strip()
        converted_date1 = convert_date(date1, full_date_input)
        date2 = date2.strip()
        converted_date2 = convert_date(date2, full_date_input)
        row_index = cell.row
        # Дублируем строку с найденной ячейкой
        sheet.insert_rows(row_index + 1)
        new_row = sheet[row_index]
        did = cell
        source_row = []
        for cell in sheet[row_index]:
            source_row.append(cell.value)
        
        # Номер строки, в которую нужно вставить скопированную строку
        

        # Вставляем скопированную строку в указанную позицию
        for mgss_sheetnames, value in enumerate(source_row, start=1):
            sheet.cell(row=row_index + 1, column=mgss_sheetnames, value=value)
        
        cell_h = sheet.cell(row=row_index, column=7)  
        value_h = cell_h.value

        if isinstance(value_h, datetime):
            new_value = value_h.replace(hour=23, minute=59, second=59)

    # Записываем новое значение в ячейку в столбце G и строке row_index
            cell_g = sheet.cell(row=row_index, column=8)  
            cell_g.value = new_value
            new_value1 = value_h.replace(hour=00, minute=00, second=00, day= int(next_dayy(converted_date1)))
            
    # Записываем новое значение в ячейку в столбце G и строке row_index
            cell_g1 = sheet.cell(row=row_index + 1, column=7)  
            cell_g1.value = new_value1
        did.value = converted_date1 
        
        for i, cell in enumerate(sheet[row_index]):
            new_row[i].value = cell.value
        new_cell = sheet.cell(row=cell.row + 1, column=2)
        new_cell.value = converted_date2
        print(converted_date2)

        
wb.save("C:/Users/Никита/Desktop/тест.xlsx")


        
# Загрузка файла Excel



for cell in sheet['E']:
    if cell.value is not None and isinstance(cell.value, str) and cell.value != "Конец":  # Проверяем, что значение не пустое и является строкой
        time_only = extract_time_from_datetime(cell.value)
        cell.value = time_only  # Заменяем значение в ячейке на только время

current_value = sheet['A2'].value
row_number = 2

while True:
    # Проверяем, следующее значение больше предыдущего на 1
    next_value = sheet.cell(row=row_number+1, column=1).value
    if next_value is None or int(next_value) != int(current_value) + 1:
        break
    
    current_value = next_value
    row_number += 1

print(row_number)
# Проход по строкам и выполнение алгоритма
for row in sheet.iter_rows(min_row=2, max_row= row_number, values_only=True):  # Начинаем с 2 строки, так как 1 строка - заголовки
    time_c = convert_to_time(row[2])  # Время в столбце C
    time_h = convert_to_time(row[7])  # Время в столбце H
    time_e = convert_to_time(row[4])  # Время в столбце E
    print(time_c, time_e, time_h)
    result = (cgc(time_e) - cgc(time_c) - cgc(time_h))  # Вычитаем времена
    
    sheet.cell(row=int(row[0]) + 1, column=9, value=result)  # Сохраняем результат в столбец I
    
for row in range(row_number+1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=2).value
    if str(cell_value).count(' 00:00:00'):
        v = str(cell_value).replace(' 00:00:00', '')
        sheet.cell(row=row, column=2).value = v
# Сохраняем изменения
wb.save("C:/Users/Никита/Desktop/тест.xlsx")


# Загрузка файла Excel



# Проход по строкам файла и вычитание времени из H из времени в G
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=8):
    time_g = row[0].value
    time_h = row[1].value

    if isinstance(time_g, datetime) and isinstance(time_h, datetime):
        new_time = time_h - time_g

        row[1].offset(column=1).value = new_time

# Сохранение изменений
wb.save("C:/Users/Никита/Desktop/тест.xlsx")



# Открываем файл Excel


current_value = int(sheet['A2'].value)
row_number = 2

while True:
    # Проверяем, следующее значение больше предыдущего на 1
    next_value = int(sheet.cell(row=row_number+1, column=1).value)
    if next_value is None or next_value != current_value + 1:
        break
    
    current_value = next_value
    row_number += 1

# Выводим номер строки последнего числа, к которому прибавили 1
print(f"Номер строки последнего числа, к которому прибавили 1: {row_number}")


for row in range(row_number, sheet.max_row + 1):
    date = sheet.cell(row=row, column=2).value
    datee = sheet.cell(row=row+1, column=2).value
    value = sheet.cell(row=row, column=9).value
    value1 = sheet.cell(row=row+1, column=9).value
    dater = sheet.cell(row=row-1, column=2).value
    value3 = sheet.cell(row=row-1, column=9).value
    print(row)
    if date == datee and date != dater:
        print(date, datee)
        print(value, value1)
        summ = cgc(clean_time_string(str(value))) +  cgc(clean_time_string(str(value1)))
        sheet.cell(row=row, column=9).value = summ
        sheet.cell(row=row+1, column=9).value = summ
    elif date == datee == dater:
        print(date, datee, dater)
        print(value, value1, value3)
        summ = cgc(clean_time_string(str(value))) +  cgc(clean_time_string(str(value1))) + cgc(clean_time_string(str(value3)))
        sheet.cell(row=row, column=9).value = summ
        sheet.cell(row=row+1, column=9).value = summ
        sheet.cell(row=row-1, column=9).value = summ


wb.save("C:/Users/Никита/Desktop/тест.xlsx")

